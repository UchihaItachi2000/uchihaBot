import openpyxl

class XL:

    # Initializing
    def __init__(self,FAdd):
        self.workBook=openpyxl.load_workbook(FAdd)
        self.sheet=self.workBook["Sheet1"]
        self.add=FAdd

    # Row count function
    def getRowCount(self):
        return(self.sheet.max_row)

    # Column count function
    def getColumnCount(self):
        return(self.sheet.max_column)

    # Read data
    def readData(self,inRow,inColumn):
        return self.sheet.cell(row=inRow,column=inColumn).value

    # Write date
    def writeDate(self,inRow,inColumn,inValue):
        self.sheet.cell(row=inRow,column=inColumn).value=inValue
        self.workBook.save(self.add)

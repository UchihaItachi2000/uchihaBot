xlpath="C:\\Users\\mphpamatrasu2000\\Desktop\\uchihaBot\\test.xlsx"

import openpyxl
wb=openpyxl.load_workbook(xlpath)
sh=wb.active #wb.get_sheet_by_name("Sheet1")
r=sh.max_row
c=sh.max_column

print(r)
print(c)

for row in range(1,r+1):
    for column in range(1,c+1):
        sh.cell(row=row,column=column).value=""
